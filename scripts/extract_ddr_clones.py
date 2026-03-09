"""
Extract DNA damage response (DDR) proteins and clones from vpCells supplementary data.

Source: Reicher et al., Nature Cell Biology, 2024
        "Pooled multicolour tagging for visualizing subcellular protein dynamics"
        https://doi.org/10.1038/s41556-024-01407-w

Supplementary Table 4 (MOESM7): vpCells clones, proteins, and localization annotations.
Download: https://static-content.springer.com/esm/art%3A10.1038%2Fs41556-024-01407-w/MediaObjects/41556_2024_1407_MOESM7_ESM.xlsx
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")


# --- DDR gene sets by pathway ---

DDR_PATHWAYS = {
    "HR": {"BRCA1", "BRCA2", "BARD1", "RAD51", "RAD51B", "RAD51C", "RAD51D",
            "PALB2", "RBBP8", "EXO1", "BLM", "WRN", "DNA2", "MUS81", "ZMYM3"},
    "NHEJ": {"XRCC5", "XRCC6", "PRKDC", "DCLRE1C", "XRCC4", "LIG4", "NHEJ1"},
    "alt-EJ": {"LIG3", "POLQ"},
    "NER": {"ERCC1", "ERCC2", "ERCC3", "ERCC4", "ERCC5", "ERCC6",
            "DDB1", "DDB2", "RAD23A", "RAD23B", "XPA"},
    "MMR": {"MSH2", "MSH3", "MSH6", "MLH1", "MLH3", "PMS1", "PMS2"},
    "BER": {"OGG1", "MUTYH", "APEX1", "APEX2", "XRCC1", "PARP1", "PARP2"},
    "sensor_kinase": {"ATM", "ATR", "CHEK1", "CHEK2"},
    "damage_sensor": {"MRE11", "MRE11A", "RAD50", "NBN", "RPA1", "RPA2", "RPA3"},
    "replication": {"PCNA", "RFC1", "RFC2", "RFC3", "RFC4", "RFC5",
                    "POLD1", "POLD2", "POLD3", "POLE", "POLE2", "FEN1",
                    "CLSPN", "TOPBP1", "TIMELESS", "TIPIN"},
    "checkpoint": {"TP53", "CDK1", "CDK2", "RB1", "MDM2", "MDM4",
                   "CDKN1A", "CDKN2A", "WEE1"},
    "chromatin_remodeling": {"SMARCA4", "SMARCA2", "SMARCB1", "SMARCC1", "SMARCC2",
                             "ARID1A", "ARID1B", "ARID2", "CHD4",
                             "KAT5", "EP300", "CREBBP", "TRRAP", "EP400", "SETD2"},
    "ubiquitin_DDR": {"RAD18", "RNF8", "RNF168", "RNF20", "RNF40", "HUWE1",
                      "CUL1", "CUL3", "USP7", "USP11", "BAP1"},
    "foci_markers": {"H2AFX", "MDC1", "TP53BP1"},
    "cohesin": {"SMC1A", "SMC3", "RAD21", "STAG2", "NIPBL"},
    "Fanconi_anemia": {"FANCA", "FANCB", "FANCC", "FANCD2", "FANCE",
                       "FANCF", "FANCG", "FANCI", "FANCL", "FANCM"},
    "TLS": {"REV1", "REV3L", "POLH", "POLI", "POLK"},
    "phase_separation": {"FUS", "EWSR1", "TAF15", "NPM1", "PML", "DAXX"},
    "helicase": {"DHX9", "DDX3X", "DDX5", "DDX17", "RECQL", "RECQL4", "RECQL5"},
    "nuclear_export": {"XPO1"},
    "transcription_DDR": {"PSIP1", "DAZAP1", "MAX"},
    "apoptosis": {"CASP8", "CASP3", "CASP9", "BAX", "BCL2"},
}

ALL_DDR_GENES = set()
for genes in DDR_PATHWAYS.values():
    ALL_DDR_GENES |= genes

GENE_TO_PATHWAY = {}
for pathway, genes in DDR_PATHWAYS.items():
    for g in genes:
        GENE_TO_PATHWAY.setdefault(g, []).append(pathway)


def load_supplementary(xlsx_path: str):
    """Load vpCells supplementary table 4."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Protein annotations
    proteins = {}
    ws = wb["vpCells_proteins_annotation_loc"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1]:
            proteins[row[1]] = {
                "vpcells_loc": row[5] or "",
                "hpa_loc": row[4] or "",
            }

    # Clone list
    clones = []
    ws = wb["vpCells_clones_list"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        clones.append({
            "clone_id": row[0],
            "experiment": row[1],
            "plate": row[2],
            "well": row[3],
            "protein_GFP": row[4] or "",
            "protein_mScarlet": row[5] or "",
            "sgRNA_GFP": row[6] or "",
            "sgRNA_mScarlet": row[7] or "",
            "filtered_cells": row[8] or 0,
            "in_1065_set": bool(row[9]),
        })

    wb.close()
    return proteins, clones


def extract_ddr(proteins, clones, min_cells=0):
    """Filter to DDR-related proteins and clones."""
    all_protein_names = set(proteins.keys())
    found = ALL_DDR_GENES & all_protein_names
    missing = ALL_DDR_GENES - all_protein_names

    # Protein rows
    protein_rows = []
    for gene in sorted(found):
        annot = proteins[gene]
        pathways = ", ".join(GENE_TO_PATHWAY.get(gene, ["unknown"]))
        protein_rows.append({
            "gene": gene,
            "vpcells_localization": annot["vpcells_loc"],
            "hpa_localization": annot["hpa_loc"],
            "ddr_pathway": pathways,
        })

    # Clone rows
    clone_rows = []
    for c in clones:
        gfp_hit = c["protein_GFP"] in found
        msc_hit = c["protein_mScarlet"] in found
        if (gfp_hit or msc_hit) and c["filtered_cells"] > min_cells:
            clone_rows.append({**c, "ddr_protein_GFP": gfp_hit, "ddr_protein_mScarlet": msc_hit})

    return protein_rows, clone_rows, found, missing


def write_csvs(protein_rows, clone_rows, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(out_dir / "vpcells_ddr_proteins.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["gene", "vpcells_localization", "hpa_localization", "ddr_pathway"])
        w.writeheader()
        w.writerows(protein_rows)

    with open(out_dir / "vpcells_ddr_clones.csv", "w", newline="") as f:
        fields = ["clone_id", "experiment", "plate", "well", "protein_GFP", "protein_mScarlet",
                  "sgRNA_GFP", "sgRNA_mScarlet", "filtered_cells", "in_1065_set",
                  "ddr_protein_GFP", "ddr_protein_mScarlet"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(clone_rows)


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "41556_2024_1407_MOESM7_ESM.xlsx"
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data")

    print(f"Loading {xlsx}...")
    proteins, clones = load_supplementary(xlsx)
    print(f"  {len(proteins)} proteins, {len(clones)} clones")

    protein_rows, clone_rows, found, missing = extract_ddr(proteins, clones, min_cells=0)
    print(f"\nDDR proteins found: {len(found)}/{len(ALL_DDR_GENES)}")
    print(f"DDR clones (cells > 0): {sum(1 for c in clone_rows if c['filtered_cells'] > 0)}")
    print(f"DDR clones in validated 1065 set: {sum(1 for c in clone_rows if c['in_1065_set'])}")

    write_csvs(protein_rows, clone_rows, out)
    print(f"\nWritten to {out}/")

    print(f"\nMissing DDR genes (not in vpCells):")
    for g in sorted(missing):
        pw = ", ".join(GENE_TO_PATHWAY.get(g, []))
        print(f"  {g:<12} ({pw})")
